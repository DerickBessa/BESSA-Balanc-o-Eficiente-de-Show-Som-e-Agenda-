from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
import os

def parse_number(value):
    """Tenta transformar value em float aceitando formatos como:
       'R$ 1.234,56', '1.234,56', '1234.56', 1234, '1.234' etc.
       Retorna float ou o value original se não for possível."""
    if value is None:
        return value
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip()
    if s == "":
        return value
    s = s.replace("R$", "").replace("r$", "").replace("$", "").strip()
    if "." in s and "," in s:
        s = s.replace(".", "")
        s = s.replace(",", ".")
    else:
        if "," in s and "." not in s:
            s = s.replace(",", ".")
    s = s.replace(" ", "")
    try:
        return float(s)
    except:
        return value

def salvar_evento_excel(nome_arquivo, dados_evento):
    if isinstance(dados_evento, dict):
        dados_evento = [dados_evento]

    path_file = f"Excel/Eventos/{nome_arquivo}.xlsx"
    os.makedirs("Excel/Eventos", exist_ok=True)

    # --- Abrir arquivo existente ou criar novo ---
    if os.path.exists(path_file):
        wb = load_workbook(path_file)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Eventos"

        # --- Cabeçalho ---
        col_count = len(dados_evento[0])
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col_count)
        ws["A1"] = "Resumo de Eventos"

    # --- Estilos ---
    titulo_font = Font(name="Arial", size=28, bold=True, color="800080")
    subtitulo_font = Font(name="Arial", size=18, bold=True, color="800080")
    dados_font = Font(name="Arial", size=14)
    white_font = Font(name="Arial", size=14, color="FFFFFF")
    fill_rosa = PatternFill("solid", fgColor="FFC0CB")
    pastel_red = PatternFill("solid", fgColor="FFB3B3")   # custo
    pastel_green = PatternFill("solid", fgColor="B3FFB3") # lucro
    thin_border = Border(
        left=Side(style='thin', color="000000"),
        right=Side(style='thin', color="000000"),
        top=Side(style='thin', color="000000"),
        bottom=Side(style='thin', color="000000")
    )
    center_align = Alignment(horizontal="center", vertical="center")

    # --- Cabeçalho ---
    ws["A1"].font = titulo_font
    ws["A1"].fill = fill_rosa
    ws["A1"].alignment = center_align
    ws["A1"].border = thin_border

    # --- Subtítulos ---
    # guardamos os headers originais para referência ao ajustar largura
    headers = list(dados_evento[0].keys())
    headers_lower = [h.lower() if h else "" for h in headers]

    if ws.max_row < 2:
        for col_num, header in enumerate(headers, 1):
            header_lower = (header or "").lower()
            # renomear visualmente os subtítulos de custo e lucro para "Custo Total" / "Lucro Total"
            if "custo" in header_lower:
                display_header = "Custo Total"
            elif "lucro" in header_lower:
                display_header = "Lucro Total"
            else:
                display_header = header
            cell = ws.cell(row=2, column=col_num, value=display_header)
            cell.font = subtitulo_font
            cell.fill = fill_rosa
            cell.alignment = center_align
            cell.border = thin_border

    # --- Dados ---
    monetary_keywords = ["custo", "lucro", "entrada", "valor", "total", "receita", "preco", "preço"]
    start_row = ws.max_row + 1
    for row_num, evento in enumerate(dados_evento, start=start_row):
        for col_num, (key, value) in enumerate(evento.items(), start=1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.alignment = center_align
            cell.border = thin_border

            key_lower = (key or "").lower()
            is_monetary = any(k in key_lower for k in monetary_keywords)

            if "custo" == key_lower:
                cell.fill = pastel_red
                cell.font = white_font
                parsed = parse_number(value)
                cell.value = parsed if isinstance(parsed, (int, float)) else value
                cell.number_format = 'R$ #,##0.00'
            elif "lucro" == key_lower:
                cell.fill = pastel_green
                cell.font = Font(name="Arial", size=dados_font.size, color="000000")
                parsed = parse_number(value)
                cell.value = parsed if isinstance(parsed, (int, float)) else value
                cell.number_format = 'R$ #,##0.00'
            elif is_monetary:
                cell.font = dados_font
                parsed = parse_number(value)
                cell.value = parsed if isinstance(parsed, (int, float)) else value
                cell.number_format = 'R$ #,##0.00'
            else:
                cell.value = value
                cell.font = dados_font

    # --- Ajuste automático de largura das colunas ---
    # aplica largura base e depois aumenta largura específica para custo/lucro
    for i, col_cells in enumerate(ws.iter_cols(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)), start=1):
        max_length = 0
        max_font_size = 0
        for cell in col_cells:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
                if cell.font and cell.font.size:
                    max_font_size = max(max_font_size, cell.font.size)
        try:
            letter = ws.cell(row=2, column=i).column_letter
            base_width = max(8, max_length * (max_font_size * 0.12))
            ws.column_dimensions[letter].width = base_width
        except:
            pass

    # --- Aumento lateral específico para CUSTO e LUCRO ---
    # fator de aumento (1.4 = 40% maior). Altere se quiser outro valor.
    aumento_fator = 1.4
    for idx, header_lower in enumerate(headers_lower, start=1):
        if "custo" in header_lower or "lucro" in header_lower:
            try:
                letter = ws.cell(row=2, column=idx).column_letter
                current = ws.column_dimensions[letter].width or 8
                ws.column_dimensions[letter].width = max(8, current * aumento_fator)
            except:
                pass

    # --- Ajuste automático da altura das linhas ---
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        max_font_size = max([cell.font.size if cell.font else 14 for cell in row])
        ws.row_dimensions[row[0].row].height = max_font_size * 1.8

    # --- Salvar ---
    wb.save(path_file)
    print(f"Planilha '{path_file}' salva com sucesso!")

def editar_evento_excel(nome_arquivo):
    pasta = "Excel/Eventos"
    caminho_completo = os.path.join(pasta, nome_arquivo)

    if not os.path.exists(caminho_completo):
        print(f"Arquivo {caminho_completo} não encontrado!")
        return

    wb = load_workbook(caminho_completo)
    ws = wb.active

    # Considerando que a linha 1 é título e linha 2 são subtítulos
    primeira_linha_dados = 3

    # Mostra todas as linhas de dados com contagem começando de 1
    print("\nEventos cadastrados:")
    for i, row in enumerate(ws.iter_rows(min_row=primeira_linha_dados, values_only=True), start=1):
        print(f"{i}: {row}")

    # Escolhe qual linha editar
    linha = input("\nDigite o número da linha que deseja editar: ").strip()
    if not linha.isdigit():
        print("Entrada inválida!")
        return

    linha = int(linha)
    linha_excel = primeira_linha_dados + linha - 1  # converte para índice real no Excel

    if linha_excel > ws.max_row:
        print("Linha fora do intervalo!")
        return

    # Edita os valores das colunas
    for col in range(1, ws.max_column + 1):
        valor_antigo = ws.cell(row=linha_excel, column=col).value
        print(f"\nColuna {col} (atual: {valor_antigo})")
        confirmar = input("Deseja editar esta coluna? (S/N): ").strip().upper()
        if confirmar == 'S':
            novo_valor = input("Digite o novo valor: ").strip()
            ws.cell(row=linha_excel, column=col, value=novo_valor)
        else:
            print("Mantendo valor atual.")

    wb.save(caminho_completo)
    print(f"\nLinha {linha} editada com sucesso no arquivo {caminho_completo}!")
